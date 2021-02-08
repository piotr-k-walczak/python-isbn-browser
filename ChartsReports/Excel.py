from PyQt5.QtWidgets import QFileDialog, QDialog
from PyQt5 import QtCore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_CURRENCY_USD_SIMPLE
from StatisticsGenerator import StatisticsGenerator

FONT = Font(name='Calibri', size=12, bold=True)
FILL = PatternFill(fgColor='eeeeee')
DOUBLE_LINE = Side(border_style="double", color="000000")
BORDER = Border(bottom=DOUBLE_LINE)
VERT_BORDER = Border(right=DOUBLE_LINE)


class ExcelReportGenerator:
    def __init__(self, data_store):
        self._data_store = data_store
        self._statistics_generator = StatisticsGenerator.create_with_data(self._data_store)
        self._workbook: Workbook = None

    def generate(self):
        dialog = QFileDialog()
        dialog.setDefaultSuffix('xlsx')
        path, ok = QFileDialog.getSaveFileName(None,
                                               "Select Directory",
                                               directory="Report.xlsx",
                                               filter="Excel (*.xlsx)")
        print(path)

        if path != "":
            self.generate_to(path)


    def generate_to(self, path):
        self._workbook = Workbook()
        self.generate_book_list()
        self.generate_book_stats()
        self.generate_author_stats()
        self.generate_subject_stats()
        self.generate_time_stats()
        self._workbook.save(filename=path)

    def header_cell(self, cell, value):
        cell.value = value
        cell.font = FONT
        cell.border = BORDER
        cell.fill = FILL

    def header_vert_cell(self, cell, value):
        self.header_cell(cell, value)
        cell.border = VERT_BORDER

    def generate_book_list(self):
        sheet = self._workbook.active
        sheet.title = "Book Data"

        self.header_cell(sheet["B2"], "Id")
        self.header_cell(sheet["C2"], "Title")
        self.header_cell(sheet["D2"], "Author")
        self.header_cell(sheet["E2"], "ISBN10")
        self.header_cell(sheet["F2"], "ISBN13")
        self.header_cell(sheet["G2"], "Cover")
        self.header_cell(sheet["H2"], "Pages")
        self.header_cell(sheet["I2"], "ReleaseDate")
        self.header_cell(sheet["J2"], "Overview")
        self.header_cell(sheet["K2"], "Synopsis")
        self.header_cell(sheet["L2"], "Publisher")
        self.header_cell(sheet["M2"], "Subjects")

        for i, book in enumerate(self._data_store.books):
            sheet.cell(row=3+i, column=2).value = book.book_id
            sheet.cell(row=3+i, column=3).value = book.title
            sheet.cell(row=3+i, column=4).value = book.author
            sheet.cell(row=3+i, column=5).value = book.isbn10
            sheet.cell(row=3+i, column=6).value = book.isbn13
            sheet.cell(row=3+i, column=7).value = book.cover
            sheet.cell(row=3+i, column=8).value = book.pages
            sheet.cell(row=3+i, column=9).value = book.release_date
            sheet.cell(row=3+i, column=10).value = book.overview
            sheet.cell(row=3+i, column=11).value = book.synopsis
            sheet.cell(row=3+i, column=11).value = book.publisher
            sheet.cell(row=3+i, column=13).value = book.assigned_subjects_as_string

    def generate_book_stats(self):
        bookStats = self._statistics_generator.get_other_statistics()
        sheet = self._workbook.create_sheet("Book Stats")
        self.write_book_cover_stats(sheet, bookStats)
        self.write_book_price_stats(sheet, bookStats)

    def write_book_cover_stats(self, sheet, book_stats):
        self.header_vert_cell(sheet["B2"], "Books by Cover")
        self.header_vert_cell(sheet["B3"], "Paperback")
        self.header_vert_cell(sheet["B4"], "Hardcover")
        self.header_vert_cell(sheet["B5"], "Mass market paperback")
        paperback = book_stats.papercover
        sheet["C3"] = paperback[0]
        sheet["D3"] = paperback[1]
        sheet["D3"].number_format = FORMAT_PERCENTAGE_00
        hardcover = book_stats.hardcover
        sheet["C4"] = hardcover[0]
        sheet["D4"] = hardcover[1]
        sheet["D4"].number_format = FORMAT_PERCENTAGE_00
        mass_market_papercover = book_stats.mass_market_papercover
        sheet["C5"] = mass_market_papercover[0]
        sheet["D5"] = mass_market_papercover[1]
        sheet["D5"].number_format = FORMAT_PERCENTAGE_00

    def write_book_price_stats(self, sheet, book_stats):
        self.header_vert_cell(sheet["F2"], "Books by Price")
        self.header_vert_cell(sheet["F3"], "Below $15")
        self.header_vert_cell(sheet["F4"], "Between $15 and $30")
        self.header_vert_cell(sheet["F5"], "Above $30")
        self.header_vert_cell(sheet["F6"], "Average Price")
        self.header_vert_cell(sheet["F8"], "Most Expensive Book")

        b15 = book_stats.below15
        sheet["G3"] = b15[0]
        sheet["H3"] = b15[1]
        sheet["H3"].number_format = FORMAT_PERCENTAGE_00

        b15to30 = book_stats.from15to30
        sheet["G4"] = b15to30[0]
        sheet["H4"] = b15to30[1]
        sheet["H4"].number_format = FORMAT_PERCENTAGE_00

        a30 = book_stats.above30
        sheet["G5"] = a30[0]
        sheet["H5"] = a30[1]
        sheet["H5"].number_format = FORMAT_PERCENTAGE_00

        avg = book_stats.avg_price
        sheet["G6"] = avg
        sheet["G6"].number_format = FORMAT_CURRENCY_USD_SIMPLE
        most_expensive = book_stats.most_expensive

        sheet["G8"] = most_expensive.title + " by " + most_expensive.author

    def generate_subject_stats(self):
        subStats = self._statistics_generator.get_subject_statistics()
        sheet = self._workbook.create_sheet("Subject Stats")

        self.header_cell(sheet["B2"], "Id")
        self.header_cell(sheet["C2"], "Name")
        self.header_cell(sheet["D2"], "Number of books")
        self.header_cell(sheet["E2"], "Market share")
        self.header_cell(sheet["F2"], "Avg. Price")

        for i, sub in enumerate(subStats.subjects_as_list):
            sheet.cell(row=3+i, column=2).value = sub.subject_id
            sheet.cell(row=3+i, column=3).value = sub.subject_name
            sheet.cell(row=3+i, column=4).value = sub.number_of_books
            sheet.cell(row=3+i, column=5).value = sub.market_share
            sheet.cell(row=3+i, column=5).number_format = FORMAT_PERCENTAGE_00
            sheet.cell(row=3+i, column=6).value = sub.avg_price
            sheet.cell(row=3+i, column=6).number_format = FORMAT_CURRENCY_USD_SIMPLE

    def generate_time_stats(self):
        timeStats = self._statistics_generator.get_time_statistics()
        sheet = self._workbook.create_sheet("Books in time")

        self.header_cell(sheet["C2"], "Number of books")
        self.header_cell(sheet["D2"], "Avg. Price")

        for i, y in enumerate(timeStats.years_as_list):
            self.header_cell(sheet.cell(row=3 + 12 * i, column=2), y.year)
            self.header_cell(sheet.cell(row=3 + 12 * i, column=3), y.number_of_books)
            self.header_cell(sheet.cell(row=3 + 12 * i, column=4), y.avg_price)
            for j, m in enumerate(y.months_as_list):
                sheet.cell(row=4 + 12 * i + j, column=2).value = m.month
                sheet.cell(row=4 + 12 * i + j, column=3).value = m.number_of_books
                sheet.cell(row=4 + 12 * i + j, column=4).value = m.avg_price
                sheet.cell(row=4 + 12 * i + j, column=4).number_format = FORMAT_CURRENCY_USD_SIMPLE

    def generate_author_stats(self):
        authorStats = self._statistics_generator.get_author_statistics()
        sheet = self._workbook.create_sheet("Authors Stats")

        self.header_cell(sheet["B2"], "Id")
        self.header_cell(sheet["C2"], "Name")
        self.header_cell(sheet["D2"], "Bio")
        self.header_cell(sheet["E2"], "Number of books")
        self.header_cell(sheet["F2"], "Books")

        for i, author in enumerate(authorStats.authors_as_list):
            sheet.cell(row=3+i, column=2).value = author.author_id
            sheet.cell(row=3+i, column=3).value = author.name
            sheet.cell(row=3+i, column=4).value = author.bio
            sheet.cell(row=3+i, column=5).value = author.number_of_books
            sheet.cell(row=3+i, column=6).value = author.books_as_string


